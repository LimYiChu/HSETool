from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse, reverse_lazy, resolve
from django.http import HttpResponse, HttpResponseRedirect
from django.contrib.auth.forms import UserCreationForm
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from .forms import *
from UploadExcel.forms import *
from django.contrib.auth import get_user_model
import matplotlib as plt
from .businesslogic import *
from .models import *
from UploadExcel.models import *
from django.views.generic import ListView, DetailView, UpdateView,TemplateView, CreateView
#test for login required
from django.contrib.auth.decorators import login_required
from django.conf import settings
import pypdftk
from django.views.generic.base import ContextMixin
from django.views.generic.edit import FormMixin
from django.core.mail import send_mail
from .reports import *
from Trackem.settings import EMAIL_HOST_USER
from django.template.loader import render_to_string
from django.template import loader
from django.core.mail import EmailMessage
from openpyxl import Workbook
import pandas as pd
from django.utils import timezone
#import mixins
from django.views.generic.detail import SingleObjectMixin
#from .forms import UserRegisterForm
# Create your views here.

from UploadExcel.forms import *
@csrf_exempt

def register (request):
    if request.method == 'POST':
        form = UserRegisterForm(request.POST)
        
            #usernames =   form.cleaned_data.get('username')
           # messages.success(request, 'account has been created')
        if form.is_valid():
            form.save()
            return render(request, 'userT/home.html')
    else:
       #form =  UserRegisterForm()
       #form = UserCreationForm()
        return render(request, 'userT/register.html', {'form': form})

def mainDashboard (request):
    
    
    #for chart.js Html only
    labelsActionee = []
    dataActionee = []

    labelsApprover = []
    dataApprover = []
    #This contain the entire Approver list for all levels
    Approver = []

    #get workshops
    studies = blgetAllStudies()

    #get all routes
    context_allRou = getuserRoutes(request,request.user.email)
    
    #Just get Actionee and Approver Routes, tied into model managers
    Actionee_R =    context_allRou.get('Actionee_Routes')
    Approver_R =    context_allRou.get('Approver_Routes') #this is a dictionary item {1,[list of routes]}

   
    charts=[]
    apprcharts =[]
    stripCount =[]
    striplabels = []
    #loop through each workshop and get counts
    #This function just does a count using model managers , calling from businesslogic.py
    
    for eachstudy in studies:
        StudyName = eachstudy.StudyName
        
        labels=[]
        labelsapp =[]
        #countbyStudies = blActionCountbyStudies(Actionee_R,StudyName,0) - old way limited by three streams 
        
        countbyStudies, labels= blActionCountbyStudiesStream(Actionee_R,StudyName,0) # unlimited actionee
        
             
        stripCount, striplabels ,  = stripAndmatch(countbyStudies,labels)
    
        # For studies check if actually assigned to it or if count is 0 then just dont generate graph
        if stripCount != []:
            
            charts.append(showPie(stripCount,striplabels,StudyName))
        
        
        #the key starts at 1 which is good thing and hence just pass to get count from queseries
        for key, value in Approver_R.items():
            for items in value:
                labelsapp.append(items.Disipline)
        
        
        
        for QueSeries, Routes in Approver_R.items():
            
           
            #listofCount= blActionCountbyStudies(Routes,StudyName,QueSeries)
            listofCountManyApprovers,labelsapp =blActionCountbyStudiesStream(Routes,StudyName,QueSeries) #improved version multiple approvers
           

            sumoflistCount = sum(listofCountManyApprovers)
            if (sumoflistCount > 0):
                labelsApprover.append('Level'+str(QueSeries))
                dataApprover.append(sumoflistCount)
        
        #dont want a blank pie to be appended on with no data since it loops through studies
        if dataApprover != []:
            apprcharts.append(showPie(dataApprover,labelsApprover,StudyName))
            #print (dataApprover)
            #print(type (dataApprover))
            #print (labelsApprover)
            #print(type (labelsApprover))

        
        dataApprover = []
        labelsApprover =[]
        stripedCount =[]
        stripedlabels =[]
        countbyStudies = []
 
       
    #get count for all approver levels just by looping through the key
    #Not very accurate but im summing approver level actions together
    # key is already rationalised in getuserRoutes
    for key, value in Approver_R.items():
        x= blfuncActionCount(value,key)
        Approver.insert(key,x)
        labelsApprover.append('Level'+str(key))
        dataApprover.append(sum(x))
    
    #apprchart=showPie(dataApprover,labelsApprover,"Approver Charts")

    
    #Context just returns to HTML so that we can use it in the HTML page
    Context = {
        'charts' : charts,
        'apprcharts' : apprcharts,
        
        #'Approver_Count'       : Approver,
        'labelsActionee' : labelsActionee,
        'dataActionee' : dataActionee,
        'labelsApprover' : labelsApprover,
        'dataApprover' : dataApprover,
            }
    return render(request, 'userT/mainDashboard.html',Context)


def getActionDetails(request, id=None):
    Items = get_object_or_404(ActionItems,id=id)
    context = {
            "Items":Items

    }
    return render(request, "userT/detailactions.html", context)

def getuserRoutes(request,useremail):
    ApproverLevel = 5
    userZemail = useremail
    Approver_Routes = {}

    #Actionee routes is straight forward
    Actionee_Routes   =   ActionRoutes.ActioneeRo.get_myroutes(userZemail)
    
    #Optimised to get all approver levels; readjust the key to 1 instead of 0
    for ApproverLevel in range(1 , ApproverLevel+1):
       Approver_Routes [ApproverLevel]  =  ActionRoutes.ApproverRo.get_myroutes(userZemail,ApproverLevel)
    #context just another form of return
    contextRoutes = {
       'Actionee_Routes' : Actionee_Routes,
       'Approver_Routes': Approver_Routes,
    }
    
    return contextRoutes

#below view is for list of actions under actionee , 
# it returns a list of actions under object_list
class ActioneeList (ListView):
    template_name   =   'userT/actionListActionee.html'
    
    def get_queryset(self):
        userZemail = self.request.user.email
        ActioneeRoutes =   ActionRoutes.ActioneeRo.get_myroutes(userZemail)
        #actioneeItems = blfuncActioneeComDisSub(ActioneeRoutes,0) - To be deleted - this was limited to 3 streams
        ActioneeActions = blallActionsComDisSub(ActioneeRoutes,0)

        
        return ActioneeActions
    
    
class ApproverList (ListView):
    template_name   =   'userT/actionListApprover.html'
    
    def get_queryset(self):
        userZemail = self.request.user.email
        ApproverActions = []
        context_allRou = getuserRoutes(self.request,userZemail)
        Approver_R =    context_allRou.get('Approver_Routes')
        
        for key, value in Approver_R.items():
            #x = blfuncActioneeComDisSub(value,key)
            allactionItems= blallActionsComDisSub(value,key)
            ApproverActions.insert(key,allactionItems)
            
        
        return ApproverActions

class DetailActioneeItems (DetailView):
    template_name   =   'userT/actionDetailActionee.html'
    #queryset = ActionItems.objects.all()

    def get_object(self):
        id1 = self.kwargs.get("id")
        return get_object_or_404(ActionItems, id=id1)


class ApproveItemsMixin(UpdateView,ListView, SingleObjectMixin):
    #paginate_by = 20
    template_name = "userT/actionUpdateApproveAction.html"
    form_class = ApproverForm
    success_url = '/ApproverList/'


    def get(self, request, *args, **kwargs):
        #uses pk key to automatically get objext
        self.object = self.get_object(queryset=ActionItems.objects.all())
        
        return super().get(request, *args, **kwargs)
    def get_object(self,queryset=None):
        queryset=ActionItems.objects.all()
       
        #self.object = self.get_object(queryset=ActionItems.objects.all())
        #return super().get(request, *args, **kwargs)
        #X = queryset.get(id=self.kwargs['pk'])
       
        return queryset.get(id=self.kwargs['pk'])
       
    def form_valid(self,form):
        
            #if form is valid just increment q series by 1 so it goes to Approver que so it goes to next queSeries
        if (self.request.POST.get('Reject')):
                #If reject que series should be 0, but need another intermediate screen for comments
                #form.instance.QueSeries = 0
            
                #Need to do below with HTTPResponseredirect because normal reverse seems to give an str error
                #reverse simply redirects to url path so can call class RejectReason below since cant really call it from fucntion call directly
                #makes sense since really django wants to work with views coming from URL paths- simply a strutured way of doing stuff
            context = {
                        'StudyActionNo' : form.instance.StudyActionNo


                }
            return HttpResponseRedirect(reverse ('RejectComments', kwargs={'forkeyid': form.instance.id}))

        if (self.request.POST.get('Cancel')):
#             
           return HttpResponseRedirect('/ActioneeList/')

        if (self.request.POST.get('Approve')): 
                #  need another intermediate screen for approval no comments
            form.instance.QueSeries += 1
            return super().form_valid(form)

        if (self.request.POST.get('Delete')): 
                #  need another intermediate screen for approval no comments
            AttachmentID = self.request.POST.get ('filepk') # hidden file that holds ID of the attachment
            ActionItemID = form.instance.id
            Status = Attachments.mdlDeleteAttachment.mgrDeleteAttachmentbyID(AttachmentID)
            
            return redirect('ActioneeFormMixin' , pk=ActionItemID)

        if (self.request.POST.get('Next')): 
            return super().form_valid(form)

    def get_context_data(self,**kwargs):
        fk = self.kwargs.get("pk")
        context = super().get_context_data(**kwargs)
        context['Rejectcomments'] = Comments.mdlComments.mgrCommentsbyFK(fk)
        context['Approver'] = True
        return context

    def get_queryset(self):
       return self.object.attachments_set.all()

class ActioneeItemsMixin(ApproveItemsMixin):
    template_name = "userT/actionUpdateApproveAction.html"
    form_class = frmUpdateActioneeForm
    
    def get_context_data(self,**kwargs):
        fk = self.kwargs.get("pk") #its actually the id and used as foreign key
        context = super().get_context_data(**kwargs)
        context['Rejectcomments'] = Comments.mdlComments.mgrCommentsbyFK(fk)
        context['Approver'] = False

        discsub = blgetAttibutesfromID(fk)
        ApproverLevel = blgetApproverLevel()
        return context

    def get_success_url(self):
        return reverse ('multiplefiles', kwargs={'forkeyid': self.object.id})

def ContactUs (request):
    return render(request, 'userT/ContactUs.html')

class RejectReason (CreateView):
    model = Comments
    template_name = 'userT/rejectReason.html'
    form_class = frmAddRejectReason
    success_url = '/ApproverList/'

    def form_valid (self,form):
        if (self.request.POST.get('Reject')):
            ID = self.kwargs['forkeyid']
            #set using model manager since we want it back to actionee it has to be set at QueSeries=0
            ActionItems.mdlQueSeries.mgrsetQueSeries(ID,0)
            form.instance.Action_id = ID
            form.instance.Username = self.request.user.email
            return super().form_valid(form)
        if (self.request.POST.get('Cancel')):
            #cant use success url, its got assocaition with dict object, so have to use below
            
            return HttpResponseRedirect('/ApproverList/')
    def get_context_data(self, **kwargs):
        fk = self.kwargs['forkeyid']
        context = super().get_context_data(**kwargs)
        context['Rejectcomments'] = Comments.mdlComments.mgrCommentsbyFK(fk)
        return context
        
def ContactUs (request):
    return render(request, 'userT/ContactUs.html')

def multiplefiles (request, **kwargs):
  
    form_multi = frmMultipleFiles()
        
    if (request.POST.get('Upload')):
        
        ID = kwargs['forkeyid']
        #set using model manager since we want it back to actionee it has to be set at QueSeries=0
        files = request.FILES.getlist('Attachment')
        for file in files:
            #should be doing via model manager , the problem is its justa line of code
            x = Attachments.objects.create(
                Attachment=file,
                Action_id=ID,
                Username=request.user.email
            )

        ActionItems.mdlQueSeries.mgrsetQueSeries(ID,1)
        
        return HttpResponseRedirect('/ActioneeList/')
            
    if (request.POST.get('Cancel')):
            #cant use success url, its got associattion with dict object, so have to use below

             return HttpResponseRedirect('/ApproverList/')

        
    context = {
        'form_multi' : form_multi

    }

    return render(request, 'userT/multiplefiles.html',context)

def createExcelReports(request,filename,**kwargs):
    
    allfields = [f.name for f in ActionItems._meta.get_fields()] 
    del allfields[0:2] # pop the first 2 in the list since it returns the foreign key
    
    allWorkshops = ActionItems.objects.all()
    
    #excel part - using from openpyxl import Workbook
    workbook = Workbook()       
    worksheet = workbook.active
    worksheet.title = 'Action Items'
            
    columns = allfields
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
    row=[]
    for actions in allWorkshops:
           
            row_num += 1
            row=[]
            for field in allfields:
                    param = 'actions.'+ str(field)
                    row.append (eval(param))
                   
            for col_num, cell_value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = cell_value
                
    workbook.save(filename)

def rptoverallStatus(request, **kwargs):
    #this function is too messy and needs to be cleaned up
    #Function on businees logic to get data based on Queseries, Actionee and Approver levels
    #most of the data is
    openActionsQueSeries = [0,1,2,3,4,5]
    closedActionsQueSeries = [6]
    allOpenActions= blfuncgetallAction('Y', openActionsQueSeries)
    allClosedActions = blfuncgetallAction('Y', closedActionsQueSeries)
    
    #this is for overall charts
    listofOpenClosed = [allOpenActions,allClosedActions]
    labelsOpenClosed = ['Open', 'Closed']
    
    chart = showPie(listofOpenClosed,labelsOpenClosed,"Overall Action Status")
    
    #this is for disc/sub-disipline
    discsub = ActionRoutes.mdlAllDiscSub.mgr_getDiscSub()
    
    #important to separate list , reusing list will fuck it up by adding list below to this one
    listcountbyDisSub= []
    listlablesDisc =[]
    listcountbyCompany= []
    listlabelsCompany = []

    #default view
    for itemPair in discsub:
        
        listcountbyDisSub.append(blgetDiscSubActionCount ('Y',itemPair,openActionsQueSeries))
        listlablesDisc.append(str(itemPair[0]))#+"/"+str(itemPair[1]))
    
    chartChanges = showPie(listcountbyDisSub,listlablesDisc, "Open Actions by Disc/Sub-Disc")

    #if generatePdf is hit, the selection is checked and graphs generated internally
    if request.method == 'POST':

        ActionStatus = request.POST.get ('ActionStatus')
        ActionsSorton = request.POST.get ('SortOn')
        ViewExcel = request.POST.get('viewExcel')
       
        if (ViewExcel):
          
            createExcelReports(request)
            
        if ActionStatus =='Open':
            if ActionsSorton == 'Company':
                Company = ActionRoutes.mdlAllCompany.mgr_getCompanyCount()
                for items in Company:
                        listcountbyCompany.append(blgetCompanyActionCount (items,openActionsQueSeries))
                            #dont need to append list as its already in the list above
                chartChanges = showPie(listcountbyCompany,Company, "Open Actions by Company")
            if ActionsSorton == 'Discipline':
                discsub = ActionRoutes.mdlAllDiscSub.mgr_getDiscSub()
                listcountbyDisSub= []
                listlablesDisc =[]
                for itemPair in discsub:
        
                    listcountbyDisSub.append(blgetDiscSubActionCount ('Y',itemPair,openActionsQueSeries))
                    listlablesDisc.append(str(itemPair[0]))#+"/"+str(itemPair[1]))
    
                chartChanges = showPie(listcountbyDisSub,listlablesDisc, "Open Actions by Disc/Sub-Disc")

        else: #This is for closed actions if selected
            if ActionsSorton == 'Company':
                Company = ActionRoutes.mdlAllCompany.mgr_getCompanyCount()
                for items in Company:
                        listcountbyCompany.append(blgetCompanyActionCount (items,closedActionsQueSeries))
                            #dont need to append list as its already in the list above
                chartChanges = showPie(listcountbyCompany,Company, "Closed Actions by Company")

            if ActionsSorton == 'Discipline':
                
                discsub = ActionRoutes.mdlAllDiscSub.mgr_getDiscSub()
                listcountbyDisSub= []
                listlablesDisc =[]
                for itemPair in discsub:
        
                    listcountbyDisSub.append(blgetDiscSubActionCount ('Y',itemPair,closedActionsQueSeries))
                    listlablesDisc.append(str(itemPair[0]))#+"/"+str(itemPair[1]))
    
                chartChanges = showPie(listcountbyDisSub,listlablesDisc, "Closed Actions by Disc/Sub-Disc")
    context = {
            "chart":chart,
            "chartChanges":chartChanges,
            "overall":True

    }
    return render (request, 'userT/reports.html',context )

def rptdiscSlice(request, **kwargs):
    
    #Function on businees logic to get data based on Queseries, Actionee and Approver levels
    #most of the data is
    TotalCount = [0,1,2,3,4,5,6]
    OpenAccount = [0,1,2,3,4,5]
    ApproverQList = [1,2,3,4,5]
    ActioneeQlist = [0]
    Company = ActionRoutes.mdlAllCompany.mgr_getCompanyCount()
    discsub = ActionRoutes.mdlAllDiscSub.mgr_getDiscSub()

    listcountbyDisSub= []
    listlablebyDisSub =[]
    totalcountbyDisSub = []
    listofstringDiscSub =[]
    Title = "Open Actions by Discipline"
    label1 = "Open Actions"
    label2 = "Total Actions"
    generalxlabel = "By Discipline"
    
    labelActionee = "Actionee"
    labelApprover = "Approver"
    TitleActApp = "Actionee-Approver Open items"
    
    listofPairActioneeCount = []
    listofPairApproverCount = []
    discsub
    for itemPair in discsub:
        
        listcountbyDisSub.append(blgetDiscSubActionCount ('Y',itemPair,OpenAccount))
        totalcountbyDisSub.append(blgetDiscSubActionCount ('Y',itemPair,TotalCount))
        listlablebyDisSub.append(str(itemPair[0])) # to include sub disc later -- +"/"+str(itemPair[1])
        listofstringDiscSub.append(str(itemPair[0]+"/"+ itemPair[1]))

        listofPairActioneeCount.append(blgetDiscSubActionCount ('Y',itemPair,[0]))
        listofPairApproverCount.append(blgetDiscSubActionCount ('Y',itemPair,ApproverQList))

    #cleaner to do a second loop
    
    listoflist = [[]]
    

    # for itemPair in discsub:
        
    #     routesfortheDiscpline = ActionRoutes.mdlgetActioneeAppr.mgr_getActApp(itemPair)

        
        
    #     for items in routesfortheDiscpline:
        
            
    #         listofPairActioneeCount.append(blgetDiscSubActionCount ('Y',itemPair,[0]))
    #         listofPairApproverCount.append(blgetDiscSubActionCount ('Y',itemPair,ApproverQList))
    #         listofPairNameCount.append(items.Actionee)
    #         
    #         listoflist.append(listofPairNameCount)
    #         listofPairNameCount = []

    chart = showbar(listcountbyDisSub,totalcountbyDisSub,listlablebyDisSub, label1,label2,generalxlabel,Title)
    chartChanges = showbar(listofPairActioneeCount,listofPairApproverCount,listlablebyDisSub,labelActionee,labelApprover,generalxlabel,TitleActApp )
    context = {
            "chart":chart,
            "chartChanges":chartChanges,
            "discpslice" : listofstringDiscSub,
            "Company" : Company
            
            }
    return render (request, 'userT/repDisc.html', context)

def rptbyUser(request, **kwargs):
    context_allRou = getuserRoutes(request,request.user.email)
    Actionee_R =    context_allRou.get('Actionee_Routes')
    
    #This function just does a count using model managers , calling from businesslogic.py
    ActioneeCount = blfuncActionCount(Actionee_R,0)
    return render (request, 'userT/reports.html')
    
def GeneratePDF (request):
    filename = [] # for appending filename place before for loop
    if (request.POST.get('GeneratePDF')):      
        x=ActionItems.objects.all()  #the row shall not contain "." because conflicting with .pdf output(typcially in header) /previously used .filter(StudyActionNo__icontains='PSD')
        y= x.values()          
        for item in y :            
            i = item["StudyActionNo"] # specify +1 for each file so it does not overwrite one file  
            j = (i + '.pdf')  # easier to breakdown j           
            del item["id"]      
            data_dict=item       
            x = 'static/multiple.pdf'             
            out_file = 'static/media/' + j   # sending file to media folder inside static folder                                                        
            generated_pdf = pypdftk.fill_form(
                pdf_path = x,
                datas = data_dict,
                out_file = out_file,                             
            )
            filename.append(str(generated_pdf)) #can only append str   
            context={
                 'filename' : filename,
                 'table': True
            }
            print(context) # context now outside for loop                   
        return render(request, 'userT/GeneratePDF.html', context)                    
    return render(request, 'userT/GeneratePDF.html')

def ReportingTable(request):
    sub = Subscribe()
    if request.method == 'POST':
        #Msg=EmailMessage()
        sub = Subscribe(request.POST)
        subject = 'Test for sending email overview'
        message = 'A summary table should present here'
        recepient = str (sub ['Email'].value())
        Msg=EmailMessage(subject, message, EMAIL_HOST_USER, [recepient])
        Msg.content_subtype="html"
        Msg.attach_file('C:\\Users\\yh_si\\Desktop\\HSETool-1\\static\\multiple.pdf')
        Msg.send()
        context ={
          'form':sub
        }
        return render(request, 'userT/ReportingTable.html',context)
    return render (request, 'userT/ReportingTable.html', {'form':sub})

#def EmailReminder (request):
#    return render(request, 'userT/EmailReminder.html')

def EmailReminder(request):
    sub = Subscribe()
    if request.method == 'POST':
        sub = Subscribe(request.POST)
        recepient = str (sub ['Email'].value())
        
        context_allRou = getuserRoutes(request,recepient)
        Actionee_R =    context_allRou.get('Actionee_Routes')  
        ActionCount = blfuncActionCount(Actionee_R,0)
        
        totalaction=sum(ActionCount)
        
        #Msg=EmailMessage()
        sub = Subscribe(request.POST)
        subject = 'Template for Action Pending Responses'
        message = 'Clients template. Your pending responses are ' + str(totalaction) + ' actions.'
        
        Msg=EmailMessage(subject, message, EMAIL_HOST_USER, [recepient])
        Msg.content_subtype="html"
        Msg.send()
        context ={
          'form':sub
        }
        return render(request, 'userT/EmailReminder.html',context)
    return render (request, 'userT/EmailReminder.html', {'form':sub})

def EmailReminderAttachment(request):
    sub = Subscribe()
    if request.method == 'POST':
        #Msg=EmailMessage()
        sub = Subscribe(request.POST)
        subject = 'Template for sending out weekly report'
        message = 'Clients weekly report template & attachment.'
        recepient = str (sub ['Email'].value())
        Msg=EmailMessage(subject, message, EMAIL_HOST_USER, [recepient])
        Msg.content_subtype="html"
        Msg.attach_file('C:\\Users\yh_si\Desktop\HSETool-1\static\weeklyreporttemplate.pdf')
        Msg.send()
        context ={
          'form':sub
        }
        return render(request, 'userT/EmailReminder.html',context)
    return render (request, 'userT/EmailReminder.html', {'form':sub})


    

def Profile (request):
    return render(request, 'userT/Profile.html')

def repPMTExcel (request):
    
    if request.method == 'POST':
        ViewExcel = request.POST.get('viewExcel')
       
        if (ViewExcel):
          
            createExcelReports(request,"AllActions.xlsx")

    return render(request, 'userT/repPMTExcel.html')

def DisciplineBreakdown (request):
    return render(request, 'userT/DisciplineBreakdown.html')

def StickyNote(request):
    return render(request, 'userT/StickyNote.html')

def IndividualBreakdownByActions(request):
    return render(request, 'userT/IndividualBreakdownByActions.html')

def IndividualBreakdownByUsers(request):
    return render(request, 'userT/IndividualBreakdownByUsers.html')
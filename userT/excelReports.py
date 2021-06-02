from .models import *
from UploadExcel.models import *
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from .businesslogic import *
from django.conf import settings

def excelAllActions(lstAttributes,lstheaders,title,columremove=False):
    
    workbook = Workbook()       
    worksheet = workbook.active
    worksheet.title = title
    row_num = 1
    
    #formatting options
    ft = Font(bold=True)
    al = Alignment( wrap_text=True)
    bd = Border (bottom=Side(border_style="double", color='FF000000'))

    for col_numH, cell_valueH in enumerate(lstheaders,1):#that 1 just starts the index count at 1 which is required or openpyxl fails
        cell = worksheet.cell(row=row_num, column=col_numH)
        cell.value = cell_valueH
        #cell.style.alignment.wrap_text=True
        
        cell.font = ft
        cell.alignment = al
        cell.border = bd
        columnAlphapet = chr(ord('@')+col_numH) #just convert numbers to alphabet dont bother trying to figure it out
        worksheet.column_dimensions[columnAlphapet].width= 30 #- just flat out set the width for everything
    
    worksheet.row_dimensions[row_num].height= 30

    for lineitem in lstAttributes:
       
        row_num += 1
        
        for col_num, cell_value in enumerate(lineitem,1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            bd = Border (bottom=Side(border_style="thin", color='FF000000'))
            cell.alignment = al
            cell.border = bd
        
        worksheet.row_dimensions[row_num].height= 50
    if columremove:
        worksheet.delete_cols(columremove)
    return workbook

def excelCompleteReport(request):
    
    allfields = [f.name for f in ActionItems._meta.get_fields()] 
    del allfields[0:2] # pop the first 3 in the list since it returns the foreign key and ID which we dont want
    
    #remove Queseries - secrets of the trade dont want it in the Excel
    #allfields = [e for e in allfields if e not in ('QueSeriesTarget')]
    allfields.remove("QueSeriesTarget") # Removed target but left QueSeries because i want to state open or close
    allfieldwithActionee = allfields.copy() # Need to copy otherwise it mutates the list
    allfieldwithActionee.insert(0,"Actionee") # has to be at the start since the loop starts with ID
    allWorkshops = ActionItems.objects.all()
    
    #excel part - using from openpyxl import Workbook
    workbook = Workbook()       
    worksheet = workbook.active
    worksheet.title = 'All Action Items'
            
    columns = allfieldwithActionee
    row_num = 1

    ft = Font(bold=True)
    al = Alignment( wrap_text=True)
    bdHead = Border (bottom=Side(border_style="double", color='FF000000'),right=Side(border_style="thin", 
                color='FF000000'))
    bdNormal = Border (bottom=Side(border_style="thin", color='FF000000'),right=Side(border_style="thin", 
                color='FF000000'))
    #bdr = Border (right=Side(border_style="thin", color='FF000000'))
    fillcell = PatternFill(start_color='69616F',
                   end_color='69616F',
                   fill_type='solid')

    for col_num, column_title in enumerate(columns, 1): #just skips ID field in meta field , print allfields to see why
                cell = worksheet.cell(row=row_num, column=col_num)
                if (column_title.lower().find("queseries") != -1): # Remove Queseries from column header
                    column_title = "Status"
                cell.value = column_title
    
                cell.font = ft
                cell.alignment = al
                cell.border = bdHead
               
                cell.fill = fillcell
                columnAlphapet = chr(ord('@')+col_num) #just convert numbers to alphabet dont bother trying to figure it out
                worksheet.column_dimensions[columnAlphapet].width= 30
    
    worksheet.row_dimensions[row_num].height= 30
    row=[]

    for actions in allWorkshops:
           
            row_num += 1
            row=[]
            for field in allfields:
                    param = 'actions.'+ str(field)
                    actionitemvalue = eval(param)
                    
                    if (str(field).lower() == 'id'):
                        discsuborg = blgetDiscSubOrgfromID(actionitemvalue)
                        Actionee = ActionRoutes.mdlgetActioneeAppr.mgr_getactioneefromtriplet(discsuborg)
                        for items in Actionee:
                            row.append(items.get('Actionee'))

                    if (str(field).lower() == 'queseries'):

                            if actionitemvalue == 99:
                                actionitemvalue = "Closed"
                            else:
                                actionitemvalue = "Open"

                    row.append (actionitemvalue)

            for col_num, cell_value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = cell_value
                    cell.alignment = al
                    cell.border = bdNormal
                    
            worksheet.row_dimensions[row_num].height= 30
    worksheet.delete_cols(2)
    
   
    return workbook